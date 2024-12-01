from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from .models import Student, StudentBulkUpload
from .serializers import StudentSerializer, StudentBulkUploadSerializer
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import random
import string

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    def get_queryset(self):
        queryset = Student.objects.select_related('user', 'cluster').all()
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        return queryset

    def create(self, request, *args, **kwargs):
        """Create student and automatically send credentials"""
        response = super().create(request, *args, **kwargs)
        
        # Automatically send credentials after successful creation
        if response.status_code == status.HTTP_201_CREATED:
            student_id = response.data.get('student_id')
            if student_id:
                try:
                    student = Student.objects.get(student_id=student_id)
                    self.send_credentials_email(student)
                    student.credentials_sent = True
                    student.save()
                except Exception as e:
                    print(f"Warning: Failed to send credentials email: {str(e)}")
                    # Don't fail the creation if email sending fails
        
        return response
    
    def send_credentials_email(self, student):
        """Send credentials email to student"""
        subject = f'Welcome to {settings.SCHOOL_NAME if hasattr(settings, "SCHOOL_NAME") else "Our Platform"} - Your Login Credentials'
        message = f"""
Dear {student.name},

Welcome to our educational platform! Your student account has been created successfully.

Login Details:
Student ID: {student.student_id}
Username: {student.user.username}
Email: {student.email}
Password: {student.password}

Please login to the student portal using these credentials and change your password immediately for security.

Best regards,
Academic Administration Team
"""
        
        send_mail(
            subject,
            message,
            settings.EMAIL_HOST_USER,
            [student.email],
            fail_silently=False,
        )

    @action(detail=False, methods=['get'])
    def template(self, request):
        """Download Excel template for bulk upload"""
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Students Template"

        # Define headers
        headers = [
            'name', 'email', 'phone', 'roll_number', 
            'year_of_admission', 'current_semester'
        ]
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add sample data
        sample_data = [
            ['John Doe', 'john@example.com', '1234567890', 'ST001', 2024, 1],
            ['Jane Smith', 'jane@example.com', '9876543210', 'ST002', 2024, 1],
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """Bulk upload students from Excel file"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']
        
        # Validate file type
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file type. Please upload an Excel file.'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Read the Excel file
            df = pd.read_excel(file)
            
            # Validate required columns
            required_columns = ['name', 'email', 'roll_number', 'year_of_admission', 'current_semester']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Track upload
            bulk_upload = StudentBulkUpload.objects.create(
                filename=file.name,
                total_records=len(df),
                uploaded_by=request.user if request.user.is_authenticated else None
            )

            created_students = []
            errors = []

            for index, row in df.iterrows():
                try:
                    # Generate random password
                    password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                    
                    student_data = {
                        'name': row['name'],
                        'email': row['email'],
                        'phone': row.get('phone', ''),
                        'roll_number': row['roll_number'],
                        'year_of_admission': int(row['year_of_admission']),
                        'current_semester': int(row['current_semester']),
                        'password': password
                    }
                    
                    serializer = StudentSerializer(data=student_data)
                    if serializer.is_valid():
                        student = serializer.save()
                        created_students.append({
                            'name': student.name,
                            'email': student.email,
                            'roll_number': student.roll_number
                        })
                    else:
                        errors.append(f"Row {index + 2}: {serializer.errors}")
                        
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")

            # Update upload record
            bulk_upload.successful_records = len(created_students)
            bulk_upload.failed_records = len(errors)
            bulk_upload.status = 'completed' if not errors else 'partial'
            bulk_upload.save()

            return Response({
                'message': f'Successfully created {len(created_students)} students',
                'created_students': created_students,
                'errors': errors,
                'upload_id': bulk_upload.id
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Error processing file: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def send_credentials(self, request, pk=None):
        """Send login credentials to student via email"""
        student = self.get_object()
        
        try:
            self.send_credentials_email(student)
            student.credentials_sent = True
            student.save()
            
            return Response({'message': 'Credentials sent successfully'})
        except Exception as e:
            return Response(
                {'error': f'Failed to send email: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Reset student password"""
        student = self.get_object()
        
        # Generate new password
        new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        student.password = new_password
        student.user.set_password(new_password)
        student.user.save()
        student.save()
        
        return Response({
            'message': 'Password reset successfully',
            'new_password': new_password
        })

    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        """Toggle student active status"""
        student = self.get_object()
        student.is_active = not student.is_active
        student.save()
        
        serializer = self.get_serializer(student)
        return Response(serializer.data)
